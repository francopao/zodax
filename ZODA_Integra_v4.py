"""
================================================================================
  ZODA - MONITOR DE VALORIZACIÓN AFP INTEGRA
  Versión: 2.1  |  Actualizado para Python 3.10+ / 2026
  Autor original: equipo Mesa de Inversiones AFP Integra
  Refactorizado por: Ingeniería de Sistemas
================================================================================

¿QUÉ HACE ESTE SCRIPT?
-----------------------
  1. Lee la base de datos de DQ_IN (posiciones del fondo) desde QEST_BD.xlsm
     y la actualiza en ZODA.xlsm.
  2. Lee datos financieros de Capital IQ (CIQ) y Bloomberg (BBG) desde ZODA.xlsm.
  3. Calcula múltiplos de valorización: P/E, P/S, P/B, EV/EBITDA (trailing y forward).
  4. Calcula Z-scores históricos de esos múltiplos para cada empresa y sus baskets.
  5. Exporta los resultados a Excel: BD en ZODA.xlsm y un archivo Multiplos_hist.xlsx.
  6. Genera un PDF con gráficos de evolución de Z-scores y múltiplos por empresa.
  7. Exporta múltiplos comparables a un Excel compartido de red.

LIBRERÍAS NECESARIAS (instalar con pip si faltan):
---------------------------------------------------
  pip install pandas numpy openpyxl matplotlib

  NOTAS DE COMPATIBILIDAD (v2.1):
  - xlwings ELIMINADO: causaba crash del kernel en Spyder. Se reemplazó por
    openpyxl puro, que funciona en Spyder, PyCharm y cualquier entorno.
  - xlrd ELIMINADO: obsoleto para .xlsx/.xlsm (solo servía para .xls legacy).
  - polars / fastexcel / calamine ELIMINADOS: reemplazados por pandas + openpyxl.
  - win32com ELIMINADO: nunca se usaba directamente en el código.
  - pandas ArrowStringArray: se fuerza dtype=object al leer Excel para evitar
    el error "'ArrowStringArray' object has no attribute 'flatten'".

ESTRUCTURA DE ARCHIVOS (todos en la carpeta 'shortcut' definida en RUTA_BASE):
-------------------------------------------------------------------------------
  ZODA.xlsm      → Archivo principal con datos BBG/CIQ y configuración
  QEST_BD.xlsm   → Base de datos de posiciones (DQ_IN)

ARCHIVOS DE SALIDA:
-------------------
  ZODA.xlsm               → Hojas "BD" y "DQ_IN" actualizadas
  Multiplos_hist.xlsx     → Histórico de múltiplos por sector GICS
  PDFs/Evolucion Z y Multiplos DD.MM.AA.pdf  → Gráficos por empresa
  Evolución de Múltiplos Comparables.xlsx    → Múltiplos para presentaciones (red)
================================================================================
"""

# ==============================================================================
# SECCIÓN 1: IMPORTACIÓN DE LIBRERÍAS
# ==============================================================================
# Librerías estándar de Python (ya vienen incluidas, no hay que instalar nada)
import os
import datetime

# Librerías de análisis de datos — instalar con: pip install pandas numpy
import numpy  as np
import pandas as pd

# Para leer y escribir archivos Excel .xlsx / .xlsm — pip install openpyxl
from openpyxl import load_workbook

# Para generar gráficos — pip install matplotlib
import matplotlib.pyplot as plt
import matplotlib.dates  as mdates
from matplotlib.backends.backend_pdf import PdfPages


# ==============================================================================
# SECCIÓN 2: CONFIGURACIÓN GENERAL
# ==============================================================================

# Marca de tiempo al inicio (para medir cuánto tarda el script)
INICIO = datetime.datetime.now()
FECHA_HOY = datetime.datetime.today().strftime('%d.%m.%y')

# ------------------------------------------------------------------------------
# RUTAS DE ARCHIVOS
# Modifica solo RUTA_BASE si cambias la ubicación de la carpeta.
# El resto se construye automáticamente desde esa base.
# ------------------------------------------------------------------------------
RUTA_BASE = r"C:\Users\usuario\OneDrive\Desktop\AFP INTEGRA\Docs privados\ZODA\shortcut"

RUTA_QEST              = os.path.join(RUTA_BASE, "QEST_BD.xlsm")
RUTA_ZODA              = os.path.join(RUTA_BASE, "ZODA.xlsm")
RUTA_MULTIPLOS_HIST    = os.path.join(RUTA_BASE, "Multiplos_hist.xlsx")

# Archivo de red para las presentaciones semestrales
RUTA_MULT_COMPARABLES = (
    r"\\sppeapp00023\Inversiones\Mesa de Inversiones\Bottom-Up"
    r"\5 Estrategia del Portafolio\Monitores\ZODA\Evolución de Múltiplos Comparables.xlsx"
)

# Carpeta donde se guardan los PDFs de gráficos
RUTA_PDF = (
    r"Z:\Mesa de Inversiones\Bottom-Up\5 Estrategia del Portafolio"
    r"\Monitores\ZODA\PDFs"
)


# ==============================================================================
# SECCIÓN 3: PARÁMETROS DEL MODELO Z-SCORE
# ==============================================================================

# Ventanas temporales para el Z-score rolling (en años)
ANIOS_Z = [1, 3, 5, 10]

# Peso de cada ventana en el Z-score compuesto (deben sumar 1)
# Orden: [1 año, 3 años, 5 años, 10 años]
PESOS_Z = [0.33, 0.33, 0.33, 0.00]

# Pesos por múltiplo para el Z-score final de cada empresa
# Orden de columnas: Trail P/E, Fwd P/E, Trail P/S, Fwd P/S,
#                    Trail P/B, Fwd P/B, Trail EV/EBITDA, Fwd EV/EBITDA
PESOS_MULT_FINANCIERAS    = [0.33, 0.00, 0.33, 0.00, 0.33, 0.00, 0.00, 0.00]
PESOS_MULT_NO_FINANCIERAS = [0.25, 0.00, 0.25, 0.00, 0.25, 0.00, 0.25, 0.00]

# Índices bursátiles de referencia
TODAS_INDICES = [
    "SPBLPGPT Index",   # S&P BVL Perú General
    "IPSA Index",        # Chile
    "IBOV Index",        # Brasil
    "COLCAP Index",      # Colombia
    "MEXBOL Index",      # México
    "MXLA Index",        # MSCI Latinoamérica (se usa como denominador en Z relativo)
]


# ==============================================================================
# SECCIÓN 4: DEFINICIÓN DE ACTIVOS (ISINs para DQ_IN)
# ==============================================================================

ISINS_DQ_IN = [
    "Fondo", "Dato", "Fecha", "SBS",
    "302012200001", "300013200002", "300013200001", "300011100001",
    "300009100001", "302002100001", "302051100001", "312051100001",
    "3220512Q1094", "302029100001", "3220082A1040", "302008100002",
    "302072200001", "302845100002", "395833228291", "312039100001",
    "302043100001", "302043100002", "302032100002", "302030100001",
    "302061100001", "302011100001", "312011100001", "302107100001",
    "3021012171A1", "345874294721", "312025100001", "312034100001",
    "312026100001", "3020836W5029", "3220292P2083",
]

# Solo los ISINs numéricos (sin las columnas de texto)
ISINS_NUMERICOS = [c for c in ISINS_DQ_IN if c not in ("Fondo", "Dato", "Fecha", "SBS")]


# ==============================================================================
# SECCIÓN 5: FUNCIÓN AUXILIAR — LECTURA DE EXCEL SIN ARROW STRINGS
# ==============================================================================

def leer_hoja_excel(ruta: str, hoja: str, n_filas: int = 7600,
                    con_encabezado: bool = False) -> pd.DataFrame:
    """
    Lee una hoja de un archivo .xlsm/.xlsx con pandas + openpyxl.

    Por qué dtype=object:
        Pandas moderno (2.x) puede usar Arrow como backend de strings, lo que
        devuelve columnas de tipo ArrowStringArray que no tienen .flatten() ni
        otros métodos clásicos. Forzar dtype=object garantiza numpy strings
        normales y evita ese error.

    Parámetros:
        ruta          : ruta completa al archivo Excel
        hoja          : nombre de la pestaña
        n_filas       : máximo de filas a leer
        con_encabezado: si True usa la primera fila como encabezado

    Retorna:
        DataFrame con columnas "column_0", "column_1", ... (si sin encabezado)
    """
    header = 0 if con_encabezado else None

    df = pd.read_excel(
        ruta,
        sheet_name=hoja,
        header=header,
        nrows=n_filas,
        engine="openpyxl",
        dtype=object,          # ← clave: evita ArrowStringArray
    )

    if not con_encabezado:
        df.columns = [f"column_{i}" for i in range(len(df.columns))]

    return df


# ==============================================================================
# SECCIÓN 6: LECTURA Y PROCESAMIENTO DE BASKETS
# ==============================================================================

def leer_baskets(ruta: str) -> tuple:
    """
    Lee la hoja "Baskets" del ZODA.xlsm y extrae tres grupos sectoriales:
        - Financieras
        - Con, Ind y HC  (Consumo, Industriales y Healthcare)
        - Mineria y Utilities

    Para cada grupo devuelve:
        todas_X    → lista de todos los tickers (empresa principal + comparables)
        empresas_X → lista de tickers de empresas principales
        basket_X   → dict {empresa_principal: [lista_de_comparables]}

    Retorna una tupla de 9 elementos en el orden indicado.
    """
    df = leer_hoja_excel(ruta, "Baskets", n_filas=100)
    df = df.dropna(axis=1, how="all")

    def extraer_grupo(nombre_grupo: str):
        """
        Filtra las filas del grupo, extrae empresas principales y sus comparables.
        Se usa .tolist() en lugar de .flatten() para compatibilidad con Arrow.
        """
        subset = df[df.iloc[:, 0] == nombre_grupo].copy()
        subset = subset.drop(subset.columns[0], axis=1)
        subset = subset.dropna(axis=1, how="all")

        # Aplanar todos los valores del subset (empresas principales + comparables)
        # Nota: usamos un loop explícito en lugar de .flatten() para evitar
        # el error de ArrowStringArray en pandas 2.x
        todos_valores = []
        for _, fila in subset.iterrows():
            todos_valores.extend(fila.tolist())

        todas = [
            x for x in todos_valores
            if x is not None and str(x) not in ("None", "nan") and pd.notna(x)
        ]
        todas = list(dict.fromkeys(todas))  # eliminar duplicados manteniendo orden

        # Empresas principales: solo la primera columna
        empresas = [
            x for x in subset.iloc[:, 0].tolist()
            if x is not None and str(x) not in ("None", "nan") and pd.notna(x)
        ]
        empresas = list(dict.fromkeys(empresas))

        # Construir el diccionario basket: empresa → [comparables]
        col_empresas = subset.iloc[:, 0].tolist()
        basket = {}
        for emp in empresas:
            idx = col_empresas.index(emp)
            comparables = subset.iloc[idx, 1:].tolist()
            comparables = [
                v for v in comparables
                if v is not None and str(v) not in ("None", "nan") and pd.notna(v)
            ]
            basket[emp] = list(dict.fromkeys(comparables))

        return todas, empresas, basket

    todas_finan,      emp_finan, basket_finan = extraer_grupo("Financieras")
    todas_cons_const, emp_cc,    basket_cc    = extraer_grupo("Con, Ind y HC")
    todas_min_utl,    emp_mu,    basket_mu    = extraer_grupo("Mineria y Utilities")

    return (todas_finan, emp_finan, basket_finan,
            todas_cons_const, emp_cc, basket_cc,
            todas_min_utl, emp_mu, basket_mu)


# ==============================================================================
# SECCIÓN 7: ACTUALIZACIÓN DE LA BASE DQ_IN
# ==============================================================================

def actualizar_DQ_IN(df_qest: pd.DataFrame, dq_in_bd: pd.DataFrame) -> pd.DataFrame:
    """
    Incorpora las filas nuevas de QEST_BD al historial DQ_IN de ZODA.

    Proceso:
        1. Elimina las primeras 10 filas de metadatos de QEST.
        2. Transpone y filtra solo los ISINs relevantes.
        3. Identifica desde qué fecha hay datos nuevos (después de la última
           fecha ya guardada en ZODA.xlsm).
        4. Filtra solo filas de tipo "Q" o "PxQ" (cantidad o monto operado).
        5. Concatena la base existente + las filas nuevas.

    Parámetros:
        df_qest  : DataFrame crudo de la hoja "DQ_IN" de QEST_BD.xlsm
        dq_in_bd : DataFrame actual de la hoja "DQ_IN" de ZODA.xlsm

    Retorna:
        DataFrame actualizado listo para escribir de vuelta en ZODA.xlsm
    """
    # Limpiar metadatos y transponer
    df_qest = df_qest.drop(df_qest.index[:10])
    df_qest = df_qest.transpose()
    df_qest = df_qest.rename(columns={df_qest.columns[0]: "Isins"})

    # Quedarse solo con los ISINs relevantes
    filtrada = df_qest[df_qest["Isins"].isin(ISINS_DQ_IN)].transpose()
    filtrada.columns = filtrada.iloc[0].tolist()
    filtrada = filtrada.iloc[1:]

    # Detectar la última fecha ya registrada en la base
    ultima_fecha_bd = dq_in_bd["Fecha"].dropna().iloc[-1]
    print(f"  → DQ_IN: última fecha registrada = {ultima_fecha_bd}")

 #   fechas_qest   = filtrada["Fecha"].tolist()  # .tolist() en vez de .values para Arrow
 #   indice_inicio = fechas_qest.index(ultima_fecha_bd) + 9

    fechas_qest = filtrada["Fecha"].tolist()

    # Normalizar la fecha buscada
    ultima_fecha_norm = pd.Timestamp(ultima_fecha_bd).normalize()
    
    # Construir lista paralela de fechas normalizadas, saltando valores no-fecha (ej: "TEST", None)
    fechas_norm = []
    for f in fechas_qest:
        try:
            if f is None or isinstance(f, str):
                fechas_norm.append(None)
            else:
                fechas_norm.append(pd.Timestamp(f).normalize())
        except Exception:
            fechas_norm.append(None)
    
    if ultima_fecha_norm not in fechas_norm:
        # Fechas válidas para mostrar rango en el error
        fechas_validas = [f for f in fechas_norm if f is not None]
        rango = f"{fechas_validas[0].date()} → {fechas_validas[-1].date()}" if fechas_validas else "sin fechas válidas"
        raise ValueError(
            f"La fecha {ultima_fecha_bd.date()} no se encontró en DQ_IN de QEST. "
            f"Rango disponible: {rango}"
        )
    
    indice_inicio = fechas_norm.index(ultima_fecha_norm) + 9
    
        


    # Extraer filas nuevas y filtrar por tipo de dato
    nuevas_filas = filtrada.iloc[indice_inicio:].reset_index(drop=True)
    nuevas_filas = nuevas_filas[
        nuevas_filas["Dato"].astype(str).str.contains(r"Q|PxQ", na=False)
    ]

    return pd.concat([dq_in_bd, nuevas_filas], ignore_index=True)


# ==============================================================================
# SECCIÓN 8: CÁLCULO DE FORWARDS A 12 MESES
# ==============================================================================

def calcular_forwards_12_meses(base: pd.DataFrame, nombres_cols: list) -> pd.DataFrame:
    """
    Para cada fecha del índice, busca el valor correspondiente 12 meses adelante.
    Si no existe la fecha exacta, toma la fecha anterior más cercana disponible.
    Si el horizonte supera el último dato disponible, mantiene el valor plano (0%).

    Se usa para estimar los valores "forward" de métricas financieras (NI, ventas, etc.)
    cuando no hay datos de consenso disponibles.

    Parámetros:
        base        : DataFrame con fechas como índice
        nombres_cols: columnas para las que se crea el forward (se renombran con prefijo Fwd_)

    Retorna:
        DataFrame original con las columnas forward añadidas al final
    """
    fechas    = pd.Series(base.index.tolist())
    fecha_max = fechas.max()
    mapa      = {f: i for i, f in enumerate(fechas)}   # búsqueda O(1)

    forwards = {col: [None] * len(fechas) for col in nombres_cols}

    for idx, dia in enumerate(fechas):
        objetivo = (dia + pd.DateOffset(months=12)).replace(hour=0, minute=0, second=0)

        for col in nombres_cols:
            col_idx = nombres_cols.index(col)

            if objetivo in mapa:
                valor = base.iloc[mapa[objetivo], col_idx]
            else:
                previas = fechas[fechas < objetivo]
                if not previas.empty:
                    cercana = previas.max()
                    valor   = base.iloc[mapa[cercana], col_idx]
                    if objetivo > fecha_max:
                        dias_extra = (objetivo - fecha_max).days
                        valor = valor * (1 + 0) ** (dias_extra / 365)  # crecimiento 0%
                else:
                    valor = None

            forwards[col][idx] = valor

    return pd.concat([base, pd.DataFrame(forwards, index=base.index)], axis=1)


# ==============================================================================
# SECCIÓN 9: LIMPIEZA DE DATOS POR EMPRESA
# ==============================================================================

# Strings que Bloomberg y CIQ usan cuando no hay dato → los tratamos como 0
_VACIOS_BBG = ["#N/A N/A", "NM", "(Invalid Formula Name)", "NA", "#VALUE!"]


def _limpiar_vacios(df: pd.DataFrame) -> pd.DataFrame:
    """Reemplaza strings vacíos de fuentes externas por 0."""
    for v in _VACIOS_BBG:
        df.replace(v, 0, inplace=True)
    return df


def limpiar_datos_CIQ(datos_hoja: pd.DataFrame,
                       fechas_trimestrales: pd.Series,
                       ticker: str) -> pd.DataFrame:
    """
    Extrae y limpia datos trimestrales de Capital IQ para un ticker.

    Capital IQ aporta 4 métricas fundamentales (trimestrales):
        IQ_NI           → Net Income (Utilidad neta)
        IQ_TOTAL_REV    → Ventas totales
        IQ_TOTAL_EQUITY → Patrimonio (Book Value)
        IQ_EBITDA       → EBITDA

    Luego calcula los forwards a 12 meses de cada una.
    """
    # Filtrar columnas de este ticker (la fila 4 contiene el nombre de la empresa)
    cols    = datos_hoja.loc[:, datos_hoja.iloc[4] == ticker]
    nombres = cols.iloc[2].tolist()
    datos   = cols.iloc[5:].copy()
    datos.columns = nombres

    # Las primeras 4 columnas son de CIQ
#    ciq = datos.iloc[:, :4].dropna(axis=0, how="all").fillna(0)
#    ciq.insert(0, "Fecha", fechas_trimestrales.values)
#    ciq = ciq.set_index("Fecha")
    # DESPUÉS:
    ciq = datos.iloc[:, :4].copy()
    # Asignar fechas ANTES del dropna para mantener alineación
    if len(fechas_trimestrales) == len(ciq):
        ciq.insert(0, "Fecha", fechas_trimestrales.values)
    else:
        # Si los tamaños no coinciden, alinear por los primeros N comunes
        n = min(len(fechas_trimestrales), len(ciq))
        ciq = ciq.iloc[:n].copy()
        ciq.insert(0, "Fecha", fechas_trimestrales.values[:n])
    ciq = ciq.set_index("Fecha")
    ciq = ciq.dropna(how="all").fillna(0)  # dropna DESPUÉS, ya con índice de fecha

    ciq = _limpiar_vacios(ciq)
    ciq = ciq.astype(np.float64)

    # Agregar columnas forward
    nombres_fwd = ["Fwd_IQ_NI", "Fwd_IQ_TOTAL_REV", "Fwd_IQ_TOTAL_EQUITY", "Fwd_IQ_EBITDA"]
    ciq = calcular_forwards_12_meses(ciq, nombres_fwd)

    return ciq


def limpiar_datos_BBG(datos_hoja: pd.DataFrame,
                       fechas_diarias: pd.Series,
                       ticker: str) -> pd.DataFrame:
    """
    Extrae y limpia datos diarios de Bloomberg para un ticker.

    Bloomberg aporta: precio (Px_last), Market Cap (CUR_MKT_CAP),
    Enterprise Value (CURR_ENTP_VAL) y múltiplos diarios.
    Estas columnas empiezan en la posición 4 de la hoja.
    """
    cols    = datos_hoja.loc[:, datos_hoja.iloc[4] == ticker]
    nombres = cols.iloc[2].tolist()
    datos   = cols.iloc[5:].copy()
    datos.columns = nombres

    # DESPUÉS:
    bbg = datos.iloc[:, 4:].copy()
    if len(fechas_diarias) == len(bbg):
        bbg.insert(0, "Fecha", fechas_diarias.values)
    else:
        n = min(len(fechas_diarias), len(bbg))
        bbg = bbg.iloc[:n].copy()
        bbg.insert(0, "Fecha", fechas_diarias.values[:n])
    bbg = bbg.set_index("Fecha")
    bbg = bbg.dropna(how="all").fillna(0)
    bbg = _limpiar_vacios(bbg)
    bbg = bbg.astype(np.float64)

    return bbg


def limpiar_datos_indice(datos_hoja: pd.DataFrame,
                          fechas_diarias: pd.Series,
                          nombre_indice: str) -> pd.DataFrame:
    """
    Extrae y limpia datos de un índice bursátil (SPBLPGPT, IBOV, etc.).
    Los índices tienen el nombre en la fila 3 (no en la 4 como las empresas).
    """
    cols = datos_hoja.loc[:, datos_hoja.iloc[3] == nombre_indice]
    nombres_fijos = [
        "PX_LAST", "Trail_PE", "Fwd_PE", "Trail_PS", "Fwd_PS",
        "Trail_PB", "Fwd_PB", "Trail_EVEBITDA", "Fwd_EVEBITDA"
    ]
    datos = cols.iloc[5:].copy()
    datos.columns = nombres_fijos
    datos = datos.fillna(0)
    datos.insert(0, "Fecha", fechas_diarias.values)
    datos = datos.set_index("Fecha")
    datos = _limpiar_vacios(datos)
    datos = datos.astype(np.float64)

    return datos


# ==============================================================================
# SECCIÓN 10: CÁLCULO DE MÚLTIPLOS DE VALORIZACIÓN
# ==============================================================================

# DESPUÉS:
def calcular_multiplos(ticker: str, datos_bbg: dict,
                        datos_ciq: dict) -> pd.DataFrame:
    """
    Calcula los 8 múltiplos de valorización para una empresa.

    ¿Qué es un múltiplo?
        Relaciona el precio que el mercado paga por una empresa (Market Cap o
        Enterprise Value) con una métrica fundamental (ganancias, ventas, etc.).
        Ejemplo: P/E = 15x significa que el mercado paga $15 por cada $1 de utilidad.

    Los datos CIQ son trimestrales y los BBG son diarios. Para combinarlos se
    convierten ambos a periodicidad trimestral con dt.to_period('Q').

    Múltiplos calculados:
        Trail P/E      → Market Cap / Net Income (últimos 12 meses reportados)
        Fwd P/E        → Market Cap / Net Income (estimado próximos 12 meses)
        Trail P/S      → Market Cap / Ventas (trailing)
        Fwd P/S        → Market Cap / Ventas (forward)
        Trail P/B      → Market Cap / Book Value (trailing)
        Fwd P/B        → Market Cap / Book Value (forward)
        Trail EV/EBITDA → Enterprise Value / EBITDA (trailing)
        Fwd EV/EBITDA  → Enterprise Value / EBITDA (forward)
    """

    # DESPUÉS:
    def bbg_trim(col: str) -> pd.Series:
        """Convierte una columna diaria de BBG a frecuencia trimestral (último valor del trimestre)."""
        s = datos_bbg[ticker][col].copy()
        s.index = s.index.to_period("Q")
        return s.groupby(level=0).last()  # un valor por trimestre, sin duplicados

    # DESPUÉS:
    def ciq_trim(col: str) -> pd.Series:
        """Convierte una columna trimestral de CIQ al mismo formato de periodo."""
        s = datos_ciq[ticker][col].copy()
        s.index = pd.DatetimeIndex(s.index).to_period("Q")
        return s.groupby(level=0).last()  # un valor por trimestre, sin duplicados

    # DESPUÉS:
    def ratio(num: pd.Series, den: pd.Series, nombre: str) -> pd.DataFrame:
        df = pd.DataFrame(num / den, columns=[nombre])
        # El índice ya viene correcto de num/den (período Q) — no reasignar fechas
        df.replace([np.inf, -np.inf], 0, inplace=True)
        return df

    # --- Obtener variables base ---
    mkt_cap = bbg_trim("CUR_MKT_CAP")
    ev      = bbg_trim("CURR_ENTP_VAL")

    net_income = ciq_trim("IQ_NI")
    ventas     = ciq_trim("IQ_TOTAL_REV")
    book       = ciq_trim("IQ_TOTAL_EQUITY")
    ebitda     = ciq_trim("IQ_EBITDA")

    fwd_ni     = ciq_trim("Fwd_IQ_NI")
    fwd_ventas = ciq_trim("Fwd_IQ_TOTAL_REV")
    fwd_book   = ciq_trim("Fwd_IQ_TOTAL_EQUITY")
    fwd_ebitda = ciq_trim("Fwd_IQ_EBITDA")

    # --- Calcular los 8 múltiplos ---
    # DESPUÉS:
    resultado = pd.concat([
        ratio(mkt_cap, net_income,  "Trail_PE"),
        ratio(mkt_cap, fwd_ni,      "Fwd_PE"),
        ratio(mkt_cap, ventas,      "Trail_PS"),
        ratio(mkt_cap, fwd_ventas,  "Fwd_PS"),
        ratio(mkt_cap, book,        "Trail_PB"),
        ratio(mkt_cap, fwd_book,    "Fwd_PB"),
        ratio(ev,      ebitda,      "Trail_EVEBITDA"),
        ratio(mkt_cap, fwd_ebitda,  "Fwd_EVEBITDA"),
    ], axis=1)
    
    # Convertir índice Period Q → Timestamp para compatibilidad con índices diarios
    if hasattr(resultado.index, 'to_timestamp'):
        resultado.index = resultado.index.to_timestamp()
    
    return resultado


# ==============================================================================
# SECCIÓN 11: CÁLCULO DE Z-SCORES
# ==============================================================================

def calcular_z_rolling(serie: pd.DataFrame, anios: int) -> pd.DataFrame:
    """
    Calcula el Z-score con ventana deslizante (rolling) de N años.

    ¿Qué es el Z-score?
        Mide cuántas desviaciones estándar está el valor actual
        respecto al promedio histórico de la ventana.
            Z > 0 → el activo está caro vs su historia
            Z < 0 → el activo está barato vs su historia
            Z = 0 → está en línea con su promedio histórico

    La ventana es de anios × 260 días (aproximando días hábiles por año).
    Se requiere la ventana completa (min_periods=ventana) para evitar
    Z-scores al inicio de la serie donde no hay suficiente historia.
    """
    # DESPUÉS:
    if hasattr(serie.index, 'to_timestamp'):
        serie.index = serie.index.to_timestamp()
    else:
        serie.index = pd.to_datetime(serie.index)
    ventana = anios * 260

    media  = serie.rolling(window=ventana, min_periods=ventana).mean()
    desvio = serie.rolling(window=ventana, min_periods=ventana).std()

    return (serie - media) / desvio


def calcular_z_final(multiplos_empresa: pd.DataFrame,
                      es_financiera: bool) -> pd.Series:
    """
    Combina los Z-scores de múltiples ventanas temporales en un único Z-score
    compuesto, ponderado por tipo de múltiplo.

    Pasos:
        1. Calcula el Z-score rolling para cada ventana (1, 3, 5, 10 años).
        2. Pondera cada ventana según PESOS_Z y los suma.
        3. Recorta la serie desde la primera fecha con datos válidos
           (la ventana más larga necesita más historia antes de tener valores).
        4. Pondera por múltiplo (PESOS_MULT_FINANCIERAS o NO_FINANCIERAS).
        5. Suma todos los múltiplos ponderados → Z-score escalar por fecha.

    Parámetros:
        multiplos_empresa : DataFrame con los 8 múltiplos de la empresa
        es_financiera     : True para empresas financieras (usan P/E, P/S, P/B)
                            False para no financieras (también usan EV/EBITDA)
    """
    z_por_ventana = {}
    for i, anios in enumerate(ANIOS_Z):
        z_por_ventana[anios] = calcular_z_rolling(multiplos_empresa, anios) * PESOS_Z[i]

    # Sumar ventanas
    z_acumulado = pd.DataFrame()
    for anios in ANIOS_Z:
        if z_acumulado.empty:
            z_acumulado = z_por_ventana[anios]
        else:
            z_acumulado = z_acumulado.add(z_por_ventana[anios], fill_value=0)

    # Recortar desde donde la ventana más larga con peso > 0 tiene datos
    # DESPUÉS:
    idx_max = max((i for i, p in enumerate(PESOS_Z) if p != 0), default=0)
    
    # Buscar la primera fecha válida desde la ventana más larga hacia la más corta
    primera_fecha = None
    for anios in reversed(ANIOS_Z[:idx_max + 1]):
        validas = z_por_ventana[anios].dropna(how="all")
        if not validas.empty:
            primera_fecha = validas.index[0]
            break
    
    if primera_fecha is not None:
        z_acumulado = z_acumulado[z_acumulado.index >= primera_fecha]
        

    # Ponderar por múltiplo y sumar
    pesos = PESOS_MULT_FINANCIERAS if es_financiera else PESOS_MULT_NO_FINANCIERAS
    return (z_acumulado * pesos).sum(axis=1)


# ==============================================================================
# SECCIÓN 12: CONSTRUCCIÓN DE BASKETS (PROMEDIOS PONDERADOS POR MARKET CAP)
# ==============================================================================

def calcular_basket(lista_comparables, todos_multiplos, market_cap, fechas):
    nombres_mult = [
        "Trail_PE", "Fwd_PE", "Trail_PS", "Fwd_PS",
        "Trail_PB", "Fwd_PB", "Trail_EVEBITDA", "Fwd_EVEBITDA"
    ]

    # Filtrar solo comparables que tienen datos calculados
    comparables_validos = [t for t in lista_comparables if t in todos_multiplos]
    if not comparables_validos:
        return pd.DataFrame(index=fechas, columns=nombres_mult)

    dfs = {m: pd.DataFrame(index=fechas) for m in nombres_mult}

    for ticker in comparables_validos:
        for mult in nombres_mult:
            temp = pd.DataFrame(todos_multiplos[ticker][mult])
            temp.rename(columns={mult: ticker}, inplace=True)
            dfs[mult] = pd.concat([dfs[mult], temp], axis=1, sort=False)  # sort=False silencia el warning

    # Pesos basados en Market Cap — solo comparables válidos presentes en market_cap
    mc_cols = [t for t in comparables_validos if t in market_cap.columns]
    if not mc_cols:
        return pd.DataFrame(index=fechas, columns=nombres_mult)

    mc = market_cap[mc_cols]
    pesos = mc.div(mc.sum(axis=1), axis=0)

    resultado = pd.DataFrame(index=fechas)
    for mult in nombres_mult:
        cols_disponibles = [c for c in mc_cols if c in dfs[mult].columns]
        resultado[mult] = pesos[cols_disponibles].multiply(
            dfs[mult][cols_disponibles]
        ).sum(axis=1)

    return resultado


# ==============================================================================
# SECCIÓN 13: ESCRITURA EN EXCEL (openpyxl, sin xlwings)
# ==============================================================================

def escribir_dataframe_en_excel(ruta: str, nombre_hoja: str,
                                 datos: pd.DataFrame,
                                 incluir_indice: bool = True) -> None:
    """
    Escribe un DataFrame en una hoja de un archivo Excel existente (.xlsm/.xlsx).

    Se usa openpyxl directamente (sin xlwings) para máxima compatibilidad:
        ✓ Funciona en Spyder
        ✓ Funciona en PyCharm
        ✓ Funciona sin tener Excel instalado
        ✓ No rompe el kernel

    IMPORTANTE: Esta función reemplaza el contenido de la hoja (la limpia
    primero y luego escribe). No modifica las macros ni otras hojas del archivo.

    Si el archivo es .xlsm (con macros), openpyxl lo lee y escribe
    correctamente sin borrar las macros.

    Parámetros:
        ruta          : ruta completa al archivo Excel
        nombre_hoja   : nombre de la pestaña donde se escribe
        datos         : DataFrame a escribir
        incluir_indice: si True, escribe el índice como primera columna
    """
    try:
        wb = load_workbook(ruta, keep_vba=True)   # keep_vba=True preserva las macros

        # Si la hoja existe la limpiamos; si no, la creamos
        if nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            ws.delete_rows(1, ws.max_row)   # borrar contenido previo
        else:
            ws = wb.create_sheet(nombre_hoja)

        # Preparar DataFrame: resetear índice si se quiere incluir
        if incluir_indice:
            df_export = datos.reset_index()
        else:
            df_export = datos.reset_index(drop=True)

        # Escribir encabezados
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        # Escribir datos fila por fila
        for row_idx, row in enumerate(df_export.itertuples(index=False), start=2):
            for col_idx, valor in enumerate(row, start=1):
                # Convertir tipos no serializables por openpyxl
                if isinstance(valor, (pd.Timestamp, datetime.datetime)):
                    valor = valor.to_pydatetime()
                elif isinstance(valor, (np.integer,)):
                    valor = int(valor)
                elif isinstance(valor, (np.floating,)):
                    valor = float(valor)
                elif pd.isna(valor) if not isinstance(valor, str) else False:
                    valor = None
                ws.cell(row=row_idx, column=col_idx, value=valor)

        wb.save(ruta)
        print(f"  ✓ Escrito: {os.path.basename(ruta)} → hoja '{nombre_hoja}'")

    except PermissionError:
        print(f"  ✗ Error: el archivo está abierto en Excel. Ciérralo y vuelve a correr.")
    except Exception as e:
        print(f"  ✗ Error al escribir en {ruta} [{nombre_hoja}]: {e}")


# ==============================================================================
# SECCIÓN 14: EXPORTAR HISTÓRICO DE MÚLTIPLOS POR GICS
# ==============================================================================

def exportar_multiplos_historicos(todos_multiplos: dict,
                                   ruta_zoda: str,
                                   ruta_salida: str) -> None:
    """
    Crea el archivo Multiplos_hist.xlsx con una pestaña por sector GICS.

    Proceso:
        1. Lee la hoja "GICS" de ZODA.xlsm para mapear ticker → empresa y sector.
        2. Concatena los múltiplos trailing de todas las empresas.
        3. Convierte a formato pivotado (filas=Fecha, columnas=Empresa+Múltiplo).
        4. Escribe una pestaña por sector en el Excel de salida.
    """
    print("  → Exportando histórico de múltiplos por sector GICS...")

    dic = leer_hoja_excel(ruta_zoda, "GICS", con_encabezado=True)
    dic = dic.apply(lambda c: c.str.strip() if c.dtype == "object" else c)
    mapa_emp  = dic.set_index("BBG")["Empresa"].to_dict()
    mapa_gics = dic.set_index("BBG")["GICS"].to_dict()

    concat = pd.concat(todos_multiplos, names=["Activo", "Fecha"]).reset_index()

    cols_trail = [
        c for c in concat.columns
        if c not in ("Activo", "Fecha") and not str(c).lower().startswith("fwd_")
    ]

    largo = concat.melt(id_vars=["Activo", "Fecha"],
                        value_vars=cols_trail,
                        var_name="Multiplo", value_name="Valor")
    largo["Fecha"] = pd.to_datetime(largo["Fecha"], errors="coerce")
    largo["Valor"] = pd.to_numeric(largo["Valor"], errors="coerce")
    largo.loc[largo["Valor"] == 0, "Valor"] = np.nan
    largo = largo.dropna(subset=["Valor"]).reset_index(drop=True)

    ancho = (largo
             .pivot(index="Fecha", columns=["Activo", "Multiplo"], values="Valor")
             .sort_index(axis=0).sort_index(axis=1))

    nuevas_cols = [(mapa_emp.get(a, a), m) for a, m in ancho.columns]
    sectores    = [mapa_gics.get(a, "Sin GICS") for a, _ in ancho.columns]
    ancho.columns = pd.MultiIndex.from_tuples(nuevas_cols, names=["Empresa", "Multiplo"])

    df_sec = pd.DataFrame({
        "Empresa": [c[0] for c in ancho.columns],
        "GICS":    sectores
    }).drop_duplicates()

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        for sector in sorted(df_sec["GICS"].unique()):
            empresas_sector = df_sec.loc[df_sec["GICS"] == sector, "Empresa"]
            cols_sector = [c for c in ancho.columns if c[0] in empresas_sector.values]
            ancho[cols_sector].to_excel(writer, sheet_name=str(sector)[:31])

    print(f"  ✓ Guardado: {ruta_salida}")


# ==============================================================================
# SECCIÓN 15: PREPARAR MÚLTIPLOS PARA EXCEL DE PRESENTACIONES
# ==============================================================================

def preparar_multiplos_trimestrales(todos_multiplos: dict, todos_baskets: dict,
                                     todas_empresas: list,
                                     fechas_trimestrales: pd.Series,
                                     fechas_diarias: pd.Index) -> pd.DataFrame:
    """
    Combina múltiplos de empresa y basket en un DataFrame trimestral
    para exportar al Excel de presentaciones semestrales.

    Retorna columnas con formato:
        "{ticker}_{multiplo}"          → múltiplo de la empresa
        "{ticker}_basket_{multiplo}"   → múltiplo del basket de comparables
    """
    partes        = []
    partes_basket = []

    for ticker in todas_empresas:
        if ticker in todos_multiplos:
            df = todos_multiplos[ticker].fillna(0).copy()
            df.columns = [f"{ticker}_{c}" for c in df.columns]
            partes.append(df)

        if ticker in todos_baskets:
            db = todos_baskets[ticker].fillna(0).copy()
            db.columns = [f"{ticker}_basket_{c}" for c in db.columns]
            partes_basket.append(db)

    combinado = pd.concat(partes + partes_basket, axis=1)

    # Alinear fechas diarias a trimestrales (tomar el dato del día más cercano anterior)
    fechas_d_sorted = pd.DatetimeIndex(fechas_diarias).sort_values()
    fechas_q_alineadas = [fechas_d_sorted.asof(q) for q in fechas_trimestrales]

    return combinado.reindex(fechas_q_alineadas)


# ==============================================================================
# SECCIÓN 16: GRÁFICOS DE EVOLUCIÓN Z-SCORE Y MÚLTIPLOS
# ==============================================================================

def _promedio_ultimo_anio(serie: pd.Series) -> float:
    """Promedio de los últimos ~260 datos (aprox. 1 año de datos hábiles)."""
    return serie.iloc[-260:].mean() if len(serie) >= 260 else serie.mean()


def graficar_empresa(df_z: pd.DataFrame, todos_multiplos: dict,
                      ticker: str, fecha_inicio: pd.Timestamp,
                      truncar: bool = True) -> plt.Figure:
    """
    Genera una figura con 9 subgráficos para una empresa:
        - 1 gráfico principal (ancho completo): evolución del Z-score compuesto
        - 8 gráficos menores: uno por cada múltiplo

    En cada gráfico de múltiplo:
        Línea azul  → valor histórico del múltiplo
        Línea cyan  → promedio histórico completo
        Línea gris  → promedio del último año

    Parámetros:
        df_z        : DataFrame que contiene la columna f"Z_{ticker}"
        todos_mult  : diccionario {ticker: DataFrame de múltiplos}
        ticker      : nombre de la empresa
        fecha_inicio: fecha desde la que graficar (ej. 2016-01-01)
        truncar     : si True, recorta el eje Y cuando hay outliers extremos
                      (útil para múltiplos que explotan en crisis)
    """
    mult = todos_multiplos[ticker]

    # Rango de fechas útil (ignorar ceros y NaN al inicio)
    fecha_min = mult[fecha_inicio:].replace(0, np.nan).dropna(how="all").index.min()
    fecha_max = mult.index.max()

    AZUL  = "#3e63ad"
    CYAN  = "#00cdcd"
    GRIS  = "#cfcfcf"

    def _subgrafico(ax, datos: pd.Series, titulo: str, con_leyenda: bool = False):
        ax.plot(datos, color=AZUL, linewidth=2)
        kw = {"linestyle": "--", "linewidth": 2}
        if con_leyenda:
            ax.axhline(datos.mean(),                   color=CYAN, label="Prom. histórico", **kw)
            ax.axhline(_promedio_ultimo_anio(datos),   color=GRIS, label="Prom. último año", **kw)
        else:
            ax.axhline(datos.mean(),                   color=CYAN, **kw)
            ax.axhline(_promedio_ultimo_anio(datos),   color=GRIS, **kw)
        ax.tick_params(axis="x", rotation=90)
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%b-%y"))
        ax.set_xlim([fecha_min, fecha_max])
        ax.set_title(titulo)
        if truncar and datos.std() > 20:
            ax.set_ylim([-5, 35])

    fig = plt.figure(figsize=(15, 20))
    fig.suptitle(ticker, fontsize=16, x=0.01, y=0.98, ha="left")

    # --- Gráfico principal: Z-score ---
    ax_z = fig.add_subplot(5, 2, (1, 2))
    ax_z.plot(df_z[f"Z_{ticker}"][fecha_min:].fillna(0), color=AZUL, linewidth=2)
    ax_z.tick_params(axis="x", rotation=90)
    ax_z.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    ax_z.xaxis.set_major_formatter(mdates.DateFormatter("%b-%y"))
    ax_z.set_xlim([fecha_min, fecha_max])
    ax_z.set_title("Evolución Z-Score Compuesto")

    # --- 8 subgráficos de múltiplos ---
    config = [
        (3,  "Trail_PE",       "Trail P/E"),
        (4,  "Fwd_PE",         "Forward P/E"),
        (5,  "Trail_PS",       "Trail P/S"),
        (6,  "Fwd_PS",         "Forward P/S"),
        (7,  "Trail_PB",       "Trail P/B"),
        (8,  "Fwd_PB",         "Forward P/B"),
        (9,  "Trail_EVEBITDA", "Trail EV/EBITDA"),
        (10, "Fwd_EVEBITDA",   "Forward EV/EBITDA"),
    ]

    for pos, col, titulo in config:
        ax = fig.add_subplot(5, 2, pos)
        datos = mult[col][fecha_min:].fillna(0)
        _subgrafico(ax, datos, titulo, con_leyenda=(pos == 10))

    fig.legend(loc="upper right")
    plt.tight_layout(rect=[0, 0.03, 1, 0.98])

    return fig


# ==============================================================================
# SECCIÓN 17: FLUJO PRINCIPAL
# ==============================================================================

def main():
    """
    Ejecuta el script completo en 9 pasos secuenciales.
    Cada paso imprime en consola lo que está haciendo y si tuvo éxito.
    """
    print("=" * 70)
    print("  ZODA — MONITOR DE VALORIZACIÓN AFP INTEGRA")
    print(f"  Inicio: {INICIO.strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 70)

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 1: Leer Baskets
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[1/9] Leyendo configuración de Baskets desde ZODA.xlsm...")
    (
        todas_finan,      emp_financieras,  basket_financieras,
        todas_cons_const, emp_cons_const,   basket_cons_const,
        todas_min_utl,    emp_min_utl,      basket_min_utl,
    ) = leer_baskets(RUTA_ZODA)

    todas_empresas = todas_finan + todas_cons_const + todas_min_utl
    print(f"  → {len(emp_financieras)} financieras | "
          f"{len(emp_cons_const)} Cons/Const | "
          f"{len(emp_min_utl)} Min/Utl")

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 2: Actualizar DQ_IN
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[2/9] Actualizando base DQ_IN...")
    dq_in_bd = leer_hoja_excel(RUTA_ZODA,  "DQ_IN", con_encabezado=True)
    df_qest  = leer_hoja_excel(RUTA_QEST,  "DQ_IN")

    base_dq_in = actualizar_DQ_IN(df_qest, dq_in_bd)
    base_dq_in = base_dq_in.reset_index(drop=True)
    base_dq_in.loc[3:, "Fecha"] = pd.to_datetime(
        base_dq_in.loc[3:, "Fecha"], dayfirst=True, errors="coerce"
    )
    for isin in ISINS_NUMERICOS:
        if isin in base_dq_in.columns:
            base_dq_in.loc[3:, isin] = pd.to_numeric(
                base_dq_in.loc[3:, isin], errors="coerce"
            )

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 3: Leer hojas de datos financieros
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[3/9] Leyendo datos de Bloomberg y Capital IQ desde ZODA.xlsm...")
    datos_financials = leer_hoja_excel(RUTA_ZODA, "Financials")
    datos_min_utl    = leer_hoja_excel(RUTA_ZODA, "Min y Uti")
    datos_cons_const = leer_hoja_excel(RUTA_ZODA, "Cons y Const")
    datos_indices    = leer_hoja_excel(RUTA_ZODA, "Indices")

    # Fechas trimestrales (columna de fecha de la hoja Financials)
    fechas_tri = pd.to_datetime(
        datos_financials["column_1"].iloc[5:], errors="coerce"
    ).dropna()

    # Fechas diarias: columnas con "Px_last" en fila 2
    mask_px = datos_financials.iloc[2] == "Px_last"
    fechas_diarias = pd.to_datetime(
        datos_financials.loc[5:, mask_px].iloc[:, 0], errors="coerce"
    ).dropna()

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 4: Limpiar datos y calcular múltiplos
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[4/9] Limpiando datos y calculando múltiplos de valorización...")

    # Mapa ticker → hoja de datos correspondiente
    # DESPUÉS:
    def ticker_existe_en_hoja(hoja: pd.DataFrame, ticker: str) -> bool:
        """Verifica que el ticker tiene al menos una columna en la hoja."""
        return (hoja.iloc[4] == ticker).any()
    
    mapa_hojas = {
        **{t: datos_financials for t in todas_finan     if ticker_existe_en_hoja(datos_financials, t)},
        **{t: datos_cons_const for t in todas_cons_const if ticker_existe_en_hoja(datos_cons_const, t)},
        **{t: datos_min_utl    for t in todas_min_utl    if ticker_existe_en_hoja(datos_min_utl,    t)},
    }
    
    # Informar cuántos tickers se descartaron por no tener datos en el Excel
    total_pedidos = len(todas_finan) + len(todas_cons_const) + len(todas_min_utl)
    print(f"  → {len(mapa_hojas)}/{total_pedidos} tickers con datos en el Excel")
    
    datos_ciq = {t: limpiar_datos_CIQ(h, fechas_tri,    t) for t, h in mapa_hojas.items()}
    datos_bbg = {t: limpiar_datos_BBG(h, fechas_diarias, t) for t, h in mapa_hojas.items()}


    datos_indices_limpios = {
        idx: limpiar_datos_indice(datos_indices, fechas_diarias, idx)
        for idx in TODAS_INDICES
    }

    # DESPUÉS:
    todos_mult   = {t: calcular_multiplos(t, datos_bbg, datos_ciq)
                    for t in mapa_hojas}
    mult_indices = {idx: datos_indices_limpios[idx].iloc[:, 1:] for idx in TODAS_INDICES}

    print(f"  → Múltiplos calculados para {len(todos_mult)} empresas")

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 5: Calcular Z-scores absolutos
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[5/9] Calculando Z-scores...")

    def _z_dict(lista_emp, es_fin):
        return {f"Z_{e}": pd.DataFrame(
            calcular_z_final(todos_mult[e], es_fin), columns=[f"Z_{e}"])
            for e in lista_emp}

    z_fin  = _z_dict(emp_financieras, True)
    z_cc   = _z_dict(emp_cons_const,  False)
    z_mu   = _z_dict(emp_min_utl,     False)
    z_idx  = {idx: pd.DataFrame(
        calcular_z_final(mult_indices[idx], False), columns=[idx])
        for idx in TODAS_INDICES}

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 6: Calcular baskets y Z relativos
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[6/9] Calculando Baskets y Z relativos (basket y BVL)...")

    market_cap = pd.concat(
        [pd.DataFrame(datos_bbg[t]["CUR_MKT_CAP"]).rename(columns={"CUR_MKT_CAP": t})
         for t in mapa_hojas], axis=1
    )

    b_fin = {e: calcular_basket(basket_financieras[e], todos_mult, market_cap, fechas_diarias)
             for e in emp_financieras}
    b_cc  = {e: calcular_basket(basket_cons_const[e],  todos_mult, market_cap, fechas_diarias)
             for e in emp_cons_const}
    b_mu  = {e: calcular_basket(basket_min_utl[e],     todos_mult, market_cap, fechas_diarias)
             for e in emp_min_utl}

    def _z_relativo(emp, mult_emp, mult_ref, es_fin, prefijo):
        ratio = mult_emp.div(mult_ref)
        return pd.DataFrame(calcular_z_final(ratio, es_fin), columns=[f"{prefijo}_{emp}"])

    bvl = mult_indices["SPBLPGPT Index"]

    z_fin_basket = {f"Z_Basket_{e}": _z_relativo(e, todos_mult[e], b_fin[e], True,  "Z_Basket")
                    for e in emp_financieras}
    z_cc_basket  = {f"Z_Basket_{e}": _z_relativo(e, todos_mult[e], b_cc[e],  False, "Z_Basket")
                    for e in emp_cons_const}
    z_mu_basket  = {f"Z_Basket_{e}": _z_relativo(e, todos_mult[e], b_mu[e],  False, "Z_Basket")
                    for e in emp_min_utl}

    z_fin_bvl = {f"Z_BVL_{e}": _z_relativo(e, todos_mult[e], bvl, True,  "Z_BVL")
                 for e in emp_financieras}
    z_cc_bvl  = {f"Z_BVL_{e}": _z_relativo(e, todos_mult[e], bvl, False, "Z_BVL")
                 for e in emp_cons_const}
    z_mu_bvl  = {f"Z_BVL_{e}": _z_relativo(e, todos_mult[e], bvl, False, "Z_BVL")
                 for e in emp_min_utl}

    mxla = mult_indices["MXLA Index"]
    z_idx_mxla = {}
    for idx in TODAS_INDICES:
        if idx == "MXLA Index":
            break
        z_idx_mxla[f"Z_MXLA_{idx}"] = pd.DataFrame(
            calcular_z_final(mult_indices[idx].div(mxla), False),
            columns=[f"Z_MXLA_{idx}"]
        )

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 7: Consolidar y exportar Z-scores a ZODA.xlsm hoja "BD"
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[7/9] Consolidando Z-scores y exportando a Excel...")

    bloques = [
        pd.concat(z_fin.values(),        axis=1),
        pd.concat(z_cc.values(),         axis=1),
        pd.concat(z_mu.values(),         axis=1),
        pd.concat(z_fin_basket.values(), axis=1),
        pd.concat(z_cc_basket.values(),  axis=1),
        pd.concat(z_mu_basket.values(),  axis=1),
        pd.concat(z_fin_bvl.values(),    axis=1),
        pd.concat(z_cc_bvl.values(),     axis=1),
        pd.concat(z_mu_bvl.values(),     axis=1),
        pd.concat(z_idx.values(),        axis=1),
    ]
    if z_idx_mxla:
        bloques.append(pd.concat(z_idx_mxla.values(), axis=1))

    base_final_z = pd.concat(bloques, axis=1)

    # Escribir en ZODA.xlsm
    escribir_dataframe_en_excel(RUTA_ZODA, "BD",    base_final_z)
    escribir_dataframe_en_excel(RUTA_ZODA, "DQ_IN", base_dq_in, incluir_indice=False)

    # Exportar histórico de múltiplos por GICS
    exportar_multiplos_historicos(todos_mult, RUTA_ZODA, RUTA_MULTIPLOS_HIST)

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 8: Exportar múltiplos comparables a la red
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[8/9] Exportando múltiplos comparables a la red...")

    todos_baskets    = {**b_fin, **b_cc, **b_mu}
    df_mult_trimestr = preparar_multiplos_trimestrales(
        todos_mult, todos_baskets, todas_empresas, fechas_tri, fechas_diarias
    )
    escribir_dataframe_en_excel(RUTA_MULT_COMPARABLES, "Multiplos", df_mult_trimestr)

    # ──────────────────────────────────────────────────────────────────────────
    # PASO 9: Generar PDF de gráficos
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[9/9] Generando PDF de gráficos...")

    df_z_fin = pd.concat(z_fin.values(), axis=1)
    df_z_cc  = pd.concat(z_cc.values(),  axis=1)
    df_z_mu  = pd.concat(z_mu.values(),  axis=1)

    fecha_inicio_graf = pd.Timestamp("2016-01-01")
    nombre_pdf = os.path.join(RUTA_PDF, f"Evolucion Z y Multiplos {FECHA_HOY}.pdf")

    try:
        with PdfPages(nombre_pdf) as pdf:
            for ticker in todos_mult:
                if ticker in emp_financieras:
                    df_z = df_z_fin
                elif ticker in emp_min_utl:
                    df_z = df_z_mu
                elif ticker in emp_cons_const:
                    df_z = df_z_cc
                else:
                    continue
                fig = graficar_empresa(df_z, todos_mult, ticker, fecha_inicio_graf)
                pdf.savefig(fig)
                plt.close(fig)
        print(f"  ✓ PDF guardado: {nombre_pdf}")
    except Exception as e:
        print(f"  ✗ Error al generar PDF: {e}")
        print(f"    ¿Existe la carpeta? → {RUTA_PDF}")

    # ──────────────────────────────────────────────────────────────────────────
    # FIN
    # ──────────────────────────────────────────────────────────────────────────
    fin = datetime.datetime.now()
    print("\n" + "=" * 70)
    print(f"  ✓ Completado en {fin - INICIO}")
    print(f"  Hora de término: {fin.strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 70)
    print(r"  \(._.)/  ¡Corrió con éxito!")
    print(r"     |")
    print(r"    7 7")


# ==============================================================================
# PUNTO DE ENTRADA
# ==============================================================================
if __name__ == "__main__":
    # Este bloque hace que main() solo se ejecute cuando corres el script
    # directamente. Si alguien importa este archivo desde otro script,
    # main() NO se ejecuta automáticamente.
    main()    
    
    